import sys, sqlite3
from PyQt5 import QtCore, QtGui, QtWidgets
from functools import partial
from mydesign import Ui_MainWindow
from Dialog_python import *
from about_autors import *
from spbgltu import *
from openpyxl import *

class MyWin(QtWidgets.QMainWindow):
    def __init__ (self, parent=None):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        
        self.ui.Building_Btn.clicked.connect(self.Building_Institute_btn_clicked)
        self.ui.Department_Btn.clicked.connect(self.Department_Btn_clicked)
        self.ui.Semester_Btn.clicked.connect(self.Semester_btn_clicked)
        self.ui.Groups_Btn.clicked.connect(self.Groups_Btn_clicked)
        self.ui.Teachers_Btn.clicked.connect(self.Teachers_Btn_clicked)
        self.ui.Classrooms_Btn.clicked.connect(self.Classrooms_Btn_clicked)
        self.ui.Subject_Btn.clicked.connect(self.Subject_Btn_clicked)
        self.ui.Sopost1_Btn.clicked.connect(self.sopost_teachers_subj)
        self.ui.Sopost2_Btn.clicked.connect(self.sopost_semestr_groups_subj)
        self.ui.start_button.clicked.connect(self.set_combobox)

        self.ui.Building_add.clicked.connect(partial(self.addInDatabase,add_code=1))
        self.ui.Institute_add.clicked.connect(partial(self.addInDatabase,add_code=2))
        self.ui.Department_add.clicked.connect(partial(self.addInDatabase,add_code=3))
        self.ui.Semester_add.clicked.connect(partial(self.addInDatabase,add_code=4))
        self.ui.Holiday_add.clicked.connect(partial(self.addInDatabase,add_code=5))
        self.ui.Position_add.clicked.connect(partial(self.addInDatabase,add_code=6))
        self.ui.Equipment_add.clicked.connect(partial(self.addInDatabase,add_code=7))
        self.ui.Groups_add.clicked.connect(partial(self.addInDatabase,add_code=8))
        self.ui.Teachers_add.clicked.connect(partial(self.addInDatabase,add_code=9))
        self.ui.Classrooms_add.clicked.connect(partial(self.addInDatabase,add_code=10))
        self.ui.Subject_add.clicked.connect(partial(self.addInDatabase,add_code=11))

        self.ui.Building_delete.clicked.connect(partial(self.DeleteFromDatabase,delete_code=1)) 
        self.ui.Institute_delete.clicked.connect(partial(self.DeleteFromDatabase,delete_code=2))
        self.ui.Department_delete.clicked.connect(partial(self.DeleteFromDatabase,delete_code=3))
        self.ui.Semester_delete.clicked.connect(partial(self.DeleteFromDatabase,delete_code=4))
        self.ui.Holiday_delete.clicked.connect(partial(self.DeleteFromDatabase,delete_code=5)) 
        self.ui.Position_delete.clicked.connect(partial(self.DeleteFromDatabase,delete_code=6))
        self.ui.Equipment_delete.clicked.connect(partial(self.DeleteFromDatabase,delete_code=7))
        self.ui.Groups_delete.clicked.connect(partial(self.DeleteFromDatabase,delete_code=8))
        self.ui.Teachers_delete.clicked.connect(partial(self.DeleteFromDatabase,delete_code=9))
        self.ui.Classrooms_delete.clicked.connect(partial(self.DeleteFromDatabase,delete_code=10))
        self.ui.Subject_delete.clicked.connect(partial(self.DeleteFromDatabase,delete_code=11))

        self.create_Tables()

        self.ui.Building_Table.setColumnHidden(0, True)
        self.ui.Institute_Table.setColumnHidden(0, True)
        self.ui.Institute_Table.setColumnHidden(1, True)
        self.ui.Department_Table.setColumnHidden(0, True)
        self.ui.Department_Table.setColumnHidden(5, True)
        self.ui.Semester_Table.setColumnHidden(0, True)
        self.ui.Semester_Table_2.setColumnHidden(0, True)
        self.ui.Holiday_Table.setColumnHidden(0, True)
        self.ui.Holiday_Table.setColumnHidden(3, True)
        self.ui.Position_Table.setColumnHidden(0, True)
        self.ui.Equipment_Table.setColumnHidden(0, True)
        self.ui.Groups_Table.setColumnHidden(0, True)
        self.ui.Groups_Table_2.setColumnHidden(0, True)
        self.ui.Teachers_Table.setColumnHidden(0, True)
        self.ui.Teachers_Table_2.setColumnHidden(0, True)
        self.ui.Classrooms_Table.setColumnHidden(0, True)
        self.ui.Subject_Table.setColumnHidden(0, True)
        self.ui.Subject_Table_2.setColumnHidden(0, True)
        self.ui.Subject_Table_3.setColumnHidden(0, True)
        self.ui.Teachers_subject_table.setColumnHidden(0, True)
        self.ui.Teachers_subject_table.setColumnHidden(1, True)
        self.ui.Teachers_subject_table.setColumnHidden(7, True)
        self.ui.Teachers_subject_table.setColumnHidden(8, True)
        self.ui.Teachers_subject_table.setColumnHidden(9, True)
        self.ui.Teachers_subject_table.setColumnHidden(10, True)
        self.ui.Teachers_subject_table.setColumnHidden(11, True)
        self.ui.Teachers_subject_table.setColumnHidden(12, True)
        self.ui.Teachers_subject_table.setColumnHidden(13, True)
        self.ui.Teachers_subject_table.setColumnHidden(14, True)
        self.ui.Teachers_subject_table.setColumnHidden(15, True)
        self.ui.Teachers_subject_table.setColumnHidden(18, True)
        self.ui.Teachers_subject_table.setColumnHidden(19, True)
        self.ui.Teachers_subject_table.setColumnHidden(20, True)
        self.ui.Teachers_subject_table.setColumnHidden(21, True)
        self.ui.Teachers_subject_table.setColumnHidden(22, True)
        self.ui.Teachers_subject_table.setColumnHidden(23, True)
        self.ui.Teachers_subject_table.setColumnHidden(24, True)
        self.ui.Teachers_subject_table.setColumnHidden(25, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(0, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(1, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(4, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(5, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(6, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(8, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(9, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(14, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(17, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(18, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(19, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(20, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(21, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(22, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(24, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(25, True)
        self.ui.Semester_Groups_Subject_table.setColumnHidden(25, True)
        self.ui.Done_table.setColumnHidden(0, True)
        self.ui.Done_table.setColumnHidden(4, True)
        self.ui.Done_table.setColumnHidden(7, True)
        self.ui.Done_table.setColumnHidden(9, True)
        self.ui.Done_table.setColumnHidden(11, True)
        
        self.LoadData(code=0) 
        self.check_choise() 
        self.check_choise_2()
        self.Building_Institute_btn_clicked()

        self.ui.Classrooms_comboBox_1.currentIndexChanged.connect(self.change_index_id_building)
        self.ui.Teachers_Table_2.itemDoubleClicked.connect(self.on_selection_teachers)
        self.ui.Subject_Table_2.itemDoubleClicked.connect(self.on_selection_in_subject)
        self.ui.Semester_Table_2.itemDoubleClicked.connect(self.on_selection_Semester_Table_2)
        self.ui.Groups_Table_2.itemDoubleClicked.connect(self.on_selection_Groups_Table_2)
        self.ui.Subject_Table_3.itemDoubleClicked.connect(self.on_selection_Subject_Table_3)

        self.ui.Teachers_subject_table.itemDoubleClicked.connect(partial(self.DeleteFromDatabase,delete_code=12))
        self.ui.Semester_Groups_Subject_table.itemDoubleClicked.connect(partial(self.DeleteFromDatabase,delete_code=13))

        self.ui.comboBox_depart.currentIndexChanged.connect(self.comboBox_depart_change)
        self.ui.comboBox_depart_2.currentIndexChanged.connect(self.comboBox_depart_2_change)
        self.ui.comboBox_Name.currentIndexChanged.connect(self.comboBox_Name_change)
        self.ui.comboBox_depart_3.currentIndexChanged.connect(self.comboBox_depart_3_change)
        self.ui.comboBox_type_group.currentIndexChanged.connect(self.comboBox_type_group_change)
        self.ui.comboBox_depart_4.currentIndexChanged.connect(self.comboBox_depart_4_change)
        self.ui.comboBox_teachers_done.currentIndexChanged.connect(self.comboBox_teachers_done_def)
        self.ui.comboBox_groups_done.currentIndexChanged.connect(self.comboBox_groups_done_def)
        self.ui.comboBox_class_done.currentIndexChanged.connect(self.comboBox_class_done_def)

        self.ui.start.clicked.connect(self.start)
        self.ui.Teachers_search.clicked.connect(self.search_in_teachers)
        self.ui.Classrooms_search.clicked.connect(self.search_in_class)
        self.ui.Subject_search.clicked.connect(self.search_in_subj) 
        self.ui.drop_tables.clicked.connect(self.drop_table)
        self.ui.About_python_button.clicked.connect(self.AboutPythonDef)
        self.ui.About_me_button.clicked.connect(self.Aboutautors)
        self.ui.About_SPBGLTU_button.clicked.connect(self.spbgltu)
        self.ui.Exit_button.clicked.connect(self.close) 
        self.ui.save_button.clicked.connect(self.save) 

    def create_Tables (self):
        con = sqlite3.connect("./mydatabase.db") 
        con.execute("PRAGMA foreign_keys = 1")
        cur=con.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS Building(ID_Building INTEGER PRIMARY KEY, Building_name TEXT, Building_placement TEXT)")
        cur.execute("CREATE TABLE IF NOT EXISTS Institute(ID_Institute INTEGER PRIMARY KEY, ID_Building INTEGER, Institute_Name TEXT, Building_name TEXT, Building_placement TEXT, FOREIGN KEY (ID_Building) REFERENCES Building(ID_Building) ON DELETE CASCADE)")
        cur.execute("CREATE TABLE IF NOT EXISTS Department(ID_Department INTEGER PRIMARY KEY, ID_Institute INTEGER, Department_Name TEXT, Institute_Name TEXT, Building_name TEXT, Building_placement TEXT, FOREIGN KEY (ID_Institute) REFERENCES Institute(ID_Institute) ON DELETE CASCADE)")
        cur.execute("CREATE TABLE IF NOT EXISTS Semester(ID_Semester INTEGER PRIMARY KEY, Semester_Name TEXT, Semester_Year TEXT, Semestr_Start TEXT, Semestr_End TEXT)")
        cur.execute("CREATE TABLE IF NOT EXISTS Holidays(ID_Holidays INTEGER PRIMARY KEY, Semester_Name_and_Year TEXT, Holiday_day TEXT, ID_Semester INTEGER, FOREIGN KEY (ID_Semester) REFERENCES Semester(ID_Semester) ON DELETE CASCADE)")
        cur.execute("CREATE TABLE IF NOT EXISTS Position(ID_Position INTEGER PRIMARY KEY, Position_name TEXT)")
        cur.execute("CREATE TABLE IF NOT EXISTS Equipment(ID_Equipment INTEGER PRIMARY KEY, Equipment_name TEXT)")
        cur.execute("CREATE TABLE IF NOT EXISTS Groups(ID_Groups INTEGER PRIMARY KEY, Groups_name TEXT, Groups_Type TEXT, Department_Name TEXT, Groups_Count INTEGER, Groups_Count_free INTEGER, Groups_Count_money INTEGER, Groups_Count_Inv INTEGER)")
        cur.execute("CREATE TABLE IF NOT EXISTS Teachers(ID_Teachers INTEGER PRIMARY KEY, Teachers_Name_1 TEXT, Teachers_Name_2 TEXT, Teachers_Name_3 TEXT, Department_Name TEXT, Date_birthday TEXT, Position_Name TEXT, Teachers_Level TEXT, Teachers_Level_2 TEXT, Teachers_Price TEXT, Teachers_Reading TEXT, Teachers_Can_Learn_Inv TEXT, Teachers_sovm TEXT, Teachers_Inv TEXT)")
        cur.execute("CREATE TABLE IF NOT EXISTS Classrooms(ID_Classrooms INTEGER PRIMARY KEY, ID_Building INTEGER, Building_name TEXT, Department_Name TEXT, Classrooms_Name TEXT, Classrooms_type TEXT, Classrooms_Count_Place INTEGER, ID_Equipment TEXT, Classrooms_Equipment_Count INTEGER, Classrooms_Inv TEXT, FOREIGN KEY (ID_Building) REFERENCES Building(ID_Building) ON DELETE CASCADE)")
        cur.execute("CREATE TABLE IF NOT EXISTS Subject(ID_Subject INTEGER PRIMARY KEY, Subject_Name TEXT, Subject_type TEXT, Subject_Hours INTEGER, Subject_Hours_lekcii INTEGER, Subject_Hours_prakt INTEGER, Subject_Hours_Self INTEGER, Subject_Hours_Exam INTEGER, Subject_Count_Zach_Ed REAL, Subject_Control TEXT, Department_Name TEXT)")
        cur.execute("CREATE TABLE IF NOT EXISTS Teachers_Subject(ID_Teachers_Subject INTEGER PRIMARY KEY, ID_Teachers INTEGER, Teachers_Name_1 TEXT, Teachers_Name_2 TEXT, Teachers_Name_3 TEXT, Department_Name_2 TEXT, Date_birthday TEXT, Position_Name TEXT, Teachers_Level TEXT, Teachers_Level_2 TEXT, Teachers_Price TEXT, Teachers_Reading TEXT, Teachers_Can_Learn_Inv TEXT, Teachers_sovm TEXT, Teachers_Inv TEXT, ID_Subject INTEGER, Subject_Name TEXT, Subject_type TEXT, Subject_Hours INTEGER, Subject_Hours_lekcii INTEGER, Subject_Hours_prakt INTEGER, Subject_Hours_Self INTEGER, Subject_Hours_Exam INTEGER, Subject_Count_Zach_Ed REAL, Subject_Control TEXT, Department_Name TEXT, Equipment_for_sopost TEXT, FOREIGN KEY (ID_Teachers) REFERENCES Teachers(ID_Teachers) ON DELETE CASCADE, FOREIGN KEY (ID_Subject) REFERENCES Subject(ID_Subject) ON DELETE CASCADE)")
        cur.execute("CREATE TABLE IF NOT EXISTS Semester_Groups_Subject(ID_Semester_Groups_Subject INTEGER PRIMARY KEY, ID_Semester INTEGER, Semester_Name TEXT, Semester_Year TEXT, Semestr_Start TEXT, Semestr_End TEXT, ID_Groups INTEGER, Groups_name TEXT, Groups_Type TEXT, Department_Name_2 TEXT, Groups_Count INTEGER, Groups_Count_free INTEGER, Groups_Count_money INTEGER, Groups_Count_Inv INTEGER,ID_Subject INTEGER, Subject_Name TEXT, Subject_type TEXT, Subject_Hours INTEGER, Subject_Hours_lekcii INTEGER, Subject_Hours_prakt INTEGER, Subject_Hours_Self INTEGER, Subject_Hours_Exam INTEGER, Subject_Count_Zach_Ed REAL, Subject_Control TEXT, Department_Name_3 TEXT, FOREIGN KEY (ID_Semester) REFERENCES Semester(ID_Semester) ON DELETE CASCADE, FOREIGN KEY (ID_Groups) REFERENCES Groups(ID_Groups) ON DELETE CASCADE, FOREIGN KEY (ID_Subject) REFERENCES Subject(ID_Subject) ON DELETE CASCADE)")
        cur.execute("CREATE TABLE IF NOT EXISTS Done(ID_Teachers INTEGER, Teachers_Name_1 TEXT, Teachers_Name_2 TEXT, Teachers_Name_3 TEXT, Number_of_day INTEGER, DAY_in_week TEXT, Number_of_para INTEGER, ID_Groups INTEGER, Groups_name TEXT, ID_Classrooms INTEGER, Classrooms_Name TEXT, ID_Subject INTEGER, Subject_Name TEXT, Subject_type TEXT, FOREIGN KEY (ID_Subject) REFERENCES Subject(ID_Subject) ON DELETE CASCADE, FOREIGN KEY (ID_Teachers) REFERENCES Teachers(ID_Teachers) ON DELETE CASCADE, FOREIGN KEY (ID_Groups) REFERENCES Groups(ID_Groups) ON DELETE CASCADE, FOREIGN KEY (ID_Classrooms) REFERENCES Classrooms(ID_Classrooms) ON DELETE CASCADE)")
        cur.close()
        con.close()

    def LoadData(self, code): 
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        if code==0:
            self.LoadData(code=1)
            self.LoadData(code=2)
            self.LoadData(code=3)
            self.LoadData(code=4)
            self.LoadData(code=5)
            self.LoadData(code=6)
            self.LoadData(code=7)
            self.LoadData(code=8)
            self.LoadData(code=9)
            self.LoadData(code=10)
            self.LoadData(code=11)
            self.LoadData(code=12)
            self.LoadData(code=13)
            self.LoadData(code=14)
            
        elif code==1: 
            query = "SELECT * FROM Building"
            result = con.execute(query)
            self.ui.Building_Table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Building_Table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Building_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                    self.ui.Building_Table.item(row_number, column_number).setTextAlignment(QtCore.Qt.AlignVCenter)
            self.ui.Building_Table.resizeColumnsToContents()
        elif code==2:
            query = "SELECT * FROM Institute"
            result = con.execute(query)
            self.ui.Institute_Table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Institute_Table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Institute_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Institute_Table.resizeColumnsToContents()
        elif code==3:
            query = "SELECT * FROM Department"
            result = con.execute(query)
            self.ui.Department_Table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Department_Table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Department_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Department_Table.resizeColumnsToContents()
        elif code==4:
            query = "SELECT * FROM Semester"
            result = con.execute(query)
            self.ui.Semester_Table.setRowCount(0)
            self.ui.Semester_Table_2.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Semester_Table.insertRow(row_number)
                self.ui.Semester_Table_2.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Semester_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                    self.ui.Semester_Table_2.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Semester_Table.resizeColumnsToContents()
            self.ui.Semester_Table_2.resizeColumnsToContents()
        elif code==5:
            query = "SELECT * FROM Holidays"
            result = con.execute(query)
            self.ui.Holiday_Table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Holiday_Table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Holiday_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Holiday_Table.resizeColumnsToContents()
        elif code==6:
            query = "SELECT * FROM Position"
            result = con.execute(query)
            self.ui.Position_Table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Position_Table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Position_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Position_Table.resizeColumnsToContents()
        elif code==7:
            query = "SELECT * FROM Equipment"
            result = con.execute(query)
            self.ui.Equipment_Table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Equipment_Table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Equipment_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Equipment_Table.resizeColumnsToContents()
        elif code==8:
            query = "SELECT * FROM Groups"
            result = con.execute(query)
            self.ui.Groups_Table.setRowCount(0)
            self.ui.Groups_Table_2.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Groups_Table.insertRow(row_number)
                self.ui.Groups_Table_2.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Groups_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                    self.ui.Groups_Table_2.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Groups_Table.resizeColumnsToContents()
            self.ui.Groups_Table_2.resizeColumnsToContents()
        elif code==9:
            query = "SELECT * FROM Teachers"
            result = con.execute(query)
            self.ui.Teachers_Table.setRowCount(0)
            self.ui.Teachers_Table_2.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Teachers_Table.insertRow(row_number)
                self.ui.Teachers_Table_2.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Teachers_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                    self.ui.Teachers_Table_2.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Teachers_Table.resizeColumnsToContents()
            self.ui.Teachers_Table_2.resizeColumnsToContents()
        elif code==10:
            query = "SELECT * FROM Classrooms"
            result = con.execute(query)
            self.ui.Classrooms_Table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Classrooms_Table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Classrooms_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Classrooms_Table.resizeColumnsToContents()
        elif code==11:
            query = "SELECT * FROM Subject"
            result = con.execute(query)
            self.ui.Subject_Table.setRowCount(0)
            self.ui.Subject_Table_2.setRowCount(0)
            self.ui.Subject_Table_3.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Subject_Table.insertRow(row_number)
                self.ui.Subject_Table_2.insertRow(row_number)
                self.ui.Subject_Table_3.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Subject_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                    self.ui.Subject_Table_2.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
                    self.ui.Subject_Table_3.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Subject_Table.resizeColumnsToContents()
            self.ui.Subject_Table_2.resizeColumnsToContents()
            self.ui.Subject_Table_3.resizeColumnsToContents()
        elif code==12:
            query = "SELECT * FROM Teachers_Subject"
            result = con.execute(query)
            self.ui.Teachers_subject_table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Teachers_subject_table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Teachers_subject_table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Teachers_subject_table.resizeColumnsToContents()
        elif code==13:
            query = "SELECT * FROM Semester_Groups_Subject"
            result = con.execute(query)
            self.ui.Semester_Groups_Subject_table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Semester_Groups_Subject_table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Semester_Groups_Subject_table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Semester_Groups_Subject_table.resizeColumnsToContents()
        elif code==14:
            query = "SELECT * FROM Done"
            result = con.execute(query)
            self.ui.Done_table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Done_table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Done_table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
            self.ui.Done_table.resizeColumnsToContents()
        cur.close()
        con.close()

    def addInDatabase(self, add_code):
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        if add_code==1:
            listToAdd =(self.ui.lineEdit_Building_Name.text(), self.ui.lineEdit_Building_Place.text())
            cur.execute('INSERT INTO Building VALUES(NULL,?,?)',listToAdd) 
            con.commit() 
            self.ui.lineEdit_Building_Name.clear()
            self.ui.lineEdit_Building_Place.clear()
            self.LoadData(code=1)
            self.Building_Institute_btn_clicked()
        elif add_code==2:
            indices=self.ui.comboBox_Buildings_In_institute.currentIndex()
            id_in_building=int(self.ui.Building_Table.item(indices, 0).text())
            query = ("SELECT Building_name, Building_placement FROM Building WHERE ID_Building LIKE ?")
            result = con.execute(query,(id_in_building,))
            listtoadd_demo=[]
            for row_number , row_data in enumerate(result):
                for column_number, data in enumerate(row_data):
                    listtoadd_demo.append(data)
            listToAdd=[]
            listToAdd.append(id_in_building)
            listToAdd.append(self.ui.lineEdit_Institute_Name.text())
            listToAdd.extend(listtoadd_demo)
            cur.execute('INSERT INTO Institute VALUES(NULL,?,?,?,?)',listToAdd)
            con.commit()
            self.ui.lineEdit_Institute_Name.clear()
            self.LoadData(code=2)
        elif add_code==3:
            indices=self.ui.comboBox_Institute_In_Department.currentIndex()
            id_in_institute=int(self.ui.Institute_Table.item(indices, 0).text())
            query = ("SELECT Institute_Name, Building_name, Building_placement FROM Institute WHERE ID_Institute LIKE ?")
            result = con.execute(query,(id_in_institute,))
            listtoadd_demo=[]
            for row_number , row_data in enumerate(result):
                for column_number, data in enumerate(row_data):
                    listtoadd_demo.append(data)
            listToAdd=[]
            listToAdd.append(id_in_institute)
            listToAdd.append(self.ui.lineEdit_Department.text())
            listToAdd.extend(listtoadd_demo)
            cur.execute('INSERT INTO Department VALUES(NULL,?,?,?,?,?)',listToAdd)
            con.commit() 
            self.ui.lineEdit_Department.clear()
            self.LoadData(code=3)
        elif add_code==4:
            data_in_Semester_comboBox=self.ui.Semester_comboBox.itemText(self.ui.Semester_comboBox.currentIndex())
            data_in_comboBox_year=self.ui.comboBox_year.itemText(self.ui.comboBox_year.currentIndex())
            listToAdd =(data_in_Semester_comboBox, data_in_comboBox_year, self.ui.dateEdit_start.text(), self.ui.dateEdit_end.text())
            cur.execute('INSERT INTO Semester VALUES(NULL,?,?,?,?)',listToAdd) 
            con.commit()
            self.LoadData(code=4)
            self.Semester_btn_clicked()
        elif add_code==5:
            data_in_Semester_And_Year_comboBox=self.ui.Semester_And_Year_comboBox.itemText(self.ui.Semester_And_Year_comboBox.currentIndex())
            v=1
            for m in range (0, count_do):
                now_data_in_len=a[v] 
                if data_in_Semester_And_Year_comboBox==now_data_in_len:
                    v=v-1
                    id_semestr=a[v]
                else:
                    v=v+2
            listToAdd =(data_in_Semester_And_Year_comboBox, self.ui.dateEdit_holiday.text(),id_semestr)
            cur.execute('INSERT INTO Holidays VALUES(NULL,?,?,?)',listToAdd) 
            con.commit() 
            self.LoadData(code=5)
        elif add_code==6:
            listToAdd =(self.ui.lineEdit_position.text(),)
            cur.execute('INSERT INTO Position VALUES(NULL,?)',listToAdd) 
            con.commit() 
            self.ui.lineEdit_position.clear()
            self.LoadData(code=6)
            self.reload_position_combobox()
        elif add_code==7:
            listToAdd =(self.ui.lineEdit_equipment.text(),)
            cur.execute('INSERT INTO Equipment VALUES(NULL,?)',listToAdd) 
            con.commit() 
            self.ui.lineEdit_equipment.clear()
            self.LoadData(code=7)
            self.Classrooms_Btn_clicked()
        elif add_code==8:
            data_in_Department_comboBox__in_groups=self.ui.Department_comboBox__in_groups.itemText(self.ui.Department_comboBox__in_groups.currentIndex())
            data_in_Groups_comboBox=self.ui.Groups_comboBox.itemText(self.ui.Groups_comboBox.currentIndex())
            listToAdd =(self.ui.Groups_search_line.text(), data_in_Groups_comboBox, data_in_Department_comboBox__in_groups, self.ui.Groups_spinBox_student.value(), self.ui.Groups_spinBox_student_2.value(), self.ui.Groups_spinBox_student_3.value(), self.ui.Groups_spinBox_student_inv.value())
            cur.execute('INSERT INTO Groups VALUES(NULL,?,?,?,?,?,?,?)',listToAdd) 
            con.commit() 
            self.ui.Groups_search_line.clear()
            self.LoadData(code=8)
        elif add_code==9:
            data_in_Teachers_comboBox_department=self.ui.Teachers_comboBox_department.itemText(self.ui.Teachers_comboBox_department.currentIndex())
            data_in_Teachers_comboBox_position=self.ui.Teachers_comboBox_position.itemText(self.ui.Teachers_comboBox_position.currentIndex())
            data_in_Teachers_comboBox_level=self.ui.Teachers_comboBox_level.itemText(self.ui.Teachers_comboBox_level.currentIndex())
            data_in_Teachers_comboBox_level_2=self.ui.Teachers_comboBox_level_2.itemText(self.ui.Teachers_comboBox_level_2.currentIndex())
            data_in_Teachers_comboBox_lec=self.ui.Teachers_comboBox_lec.itemText(self.ui.Teachers_comboBox_lec.currentIndex())
            data_in_Teachers_comboBox_teach_inv=self.ui.Teachers_comboBox_teach_inv.itemText(self.ui.Teachers_comboBox_teach_inv.currentIndex())
            data_in_Teachers_comboBox_sovm=self.ui.Teachers_comboBox_sovm.itemText(self.ui.Teachers_comboBox_sovm.currentIndex())
            data_in_Teachers_comboBox_inv=self.ui.Teachers_comboBox_inv.itemText(self.ui.Teachers_comboBox_inv.currentIndex())
            listToAdd =(self.ui.Teachers_Name1.text(), self.ui.Teachers_Name2.text(), self.ui.Teachers_Name3.text(), data_in_Teachers_comboBox_department, self.ui.dateEdit_birthday.text(), data_in_Teachers_comboBox_position, data_in_Teachers_comboBox_level, data_in_Teachers_comboBox_level_2, self.ui.Teachers_Stavka.value(), data_in_Teachers_comboBox_lec, data_in_Teachers_comboBox_teach_inv, data_in_Teachers_comboBox_sovm, data_in_Teachers_comboBox_inv)
            cur.execute('INSERT INTO Teachers VALUES(NULL,?,?,?,?,?,?,?,?,?,?,?,?,?)',listToAdd)
            con.commit() 
            self.ui.Teachers_Name1.clear()
            self.ui.Teachers_Name2.clear()
            self.ui.Teachers_Name3.clear()
            self.LoadData(code=9)
        elif add_code==10:
            indices=self.ui.Classrooms_comboBox_1.currentIndex()
            id_in_building=int(self.ui.Building_Table.item(indices, 0).text())
            query = ("SELECT Building_name FROM Building WHERE ID_Building LIKE ?")
            result = con.execute(query,(id_in_building,))
            listtoadd_demo=[]
            for row_number , row_data in enumerate(result):
                for column_number, data in enumerate(row_data):
                    listtoadd_demo.append(data)
            listToAdd=[]
            listToAdd.append(id_in_building)
            listToAdd.extend(listtoadd_demo)
            data_in_Classrooms_comboBox_2=self.ui.Classrooms_comboBox_2.itemText(self.ui.Classrooms_comboBox_2.currentIndex())
            data_in_Classrooms_comboBox_3=self.ui.Classrooms_comboBox_3.itemText(self.ui.Classrooms_comboBox_3.currentIndex())
            data_in_Classrooms_comboBox_type=self.ui.Classrooms_comboBox_type.itemText(self.ui.Classrooms_comboBox_type.currentIndex()) 
            data_in_Classrooms_comboBox_inv=self.ui.Classrooms_comboBox_inv.itemText(self.ui.Classrooms_comboBox_inv.currentIndex())
            listToAdd.append(data_in_Classrooms_comboBox_2)
            listToAdd.append(self.ui.Classrooms_search_line.text())
            listToAdd.append(data_in_Classrooms_comboBox_type)
            listToAdd.append(self.ui.Classrooms_spinBox_places.value())
            listToAdd.append(data_in_Classrooms_comboBox_3)
            listToAdd.append(self.ui.Classrooms_spinBox_places_2.value())
            listToAdd.append(data_in_Classrooms_comboBox_inv)
            cur.execute('INSERT INTO Classrooms VALUES(NULL,?,?,?,?,?,?,?,?,?)',listToAdd) 
            con.commit() 
            self.ui.Classrooms_search_line.clear()
            self.LoadData(code=10)
        elif add_code==11:
            data_in_Subject_comboBox_department=self.ui.Subject_comboBox_department.itemText(self.ui.Subject_comboBox_department.currentIndex())
            data_in_Subject_comboBox_1=self.ui.Subject_comboBox_1.itemText(self.ui.Subject_comboBox_1.currentIndex())
            data_in_Subject_comboBox_control=self.ui.Subject_comboBox_control.itemText(self.ui.Subject_comboBox_control.currentIndex())
            listToAdd =(self.ui.Subject_search_line.text(), data_in_Subject_comboBox_1, self.ui.Subject_spinBox_hours.value(), self.ui.Subject_spinBox_hours_2.value(), self.ui.Subject_spinBox_hours_3.value(), self.ui.Subject_spinBox_hours_self.value(), self.ui.Subject_spinBox_hours_attest.value(), self.ui.Subject_spinBox_zach_ed.value(),data_in_Subject_comboBox_control, data_in_Subject_comboBox_department)
            cur.execute('INSERT INTO Subject VALUES(NULL,?,?,?,?,?,?,?,?,?,?)',listToAdd) 
            con.commit() 
            self.ui.Subject_search_line.clear()
            self.LoadData(code=11)
        elif add_code==12:
            self.check_choise()
            global listOfVars_in_teach, listofVars_in_subj
            lists=[]
            lists.extend(listOfVars_in_teach)
            lists.extend(listofVars_in_subj)
            data_in_Equipment_for_sopost=self.ui.Equipment_for_sopost.itemText(self.ui.Equipment_for_sopost.currentIndex())
            lists.append(data_in_Equipment_for_sopost)
            cur.execute('INSERT INTO Teachers_Subject VALUES(NULL,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',lists)
            con.commit() 
            self.clear_teachers()
            self.clear_subj()
            self.LoadData(code=12)
        elif add_code==13:
            self.check_choise_2()
            global listOfVars_in_semestr, listOfVars_in_groups, listOfVars_in_subject
            lists=[]
            lists.extend(listOfVars_in_semestr)
            lists.extend(listOfVars_in_groups)
            lists.extend(listOfVars_in_subject)
            cur.execute('INSERT INTO Semester_Groups_Subject VALUES(NULL,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',lists) 
            con.commit() 
            self.clear_Semester()
            self.clear_Groups()
            self.clear_Subject()
            self.LoadData(code=13)
        cur.close() 
        con.close() 

    def DeleteFromDatabase(self, delete_code):
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        if delete_code==1:
            indices = self.ui.Building_Table.selectionModel().selectedRows()
            self.ui.Building_Table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Building_Table.removeRow(index.row())
                self.ui.Building_Table.selectionModel().clearCurrentIndex()
                self.ui.Building_Table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Building Where ID_Building = ?',(idInDb,))
                con.commit()
            self.ui.lineEdit_Building_Name.clear()
            self.ui.lineEdit_Building_Place.clear()
            self.LoadData(code=0)
            self.Building_Institute_btn_clicked()
        if delete_code==2:
            indices = self.ui.Institute_Table.selectionModel().selectedRows()
            self.ui.Institute_Table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Institute_Table.removeRow(index.row())
                self.ui.Institute_Table.selectionModel().clearCurrentIndex()
                self.ui.Institute_Table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Institute Where ID_Institute = ?',(idInDb,))
                con.commit()
            self.ui.lineEdit_Institute_Name.clear()
            self.LoadData(code=0)
        if delete_code==3:
            indices = self.ui.Department_Table.selectionModel().selectedRows()
            self.ui.Department_Table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Department_Table.removeRow(index.row())
                self.ui.Department_Table.selectionModel().clearCurrentIndex()
                self.ui.Department_Table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Department Where ID_Department = ?',(idInDb,))
                con.commit()
            self.LoadData(code=0)
        if delete_code==4:
            indices = self.ui.Semester_Table.selectionModel().selectedRows()
            self.ui.Semester_Table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Semester_Table.removeRow(index.row())
                self.ui.Semester_Table.selectionModel().clearCurrentIndex()
                self.ui.Semester_Table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Semester Where ID_Semester = ?',(idInDb,))
                con.commit()
            self.LoadData(code=0)
            self.Semester_btn_clicked()
        if delete_code==5:
            indices = self.ui.Holiday_Table.selectionModel().selectedRows()
            self.ui.Holiday_Table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Holiday_Table.removeRow(index.row())
                self.ui.Holiday_Table.selectionModel().clearCurrentIndex()
                self.ui.Holiday_Table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Holidays Where ID_Holidays = ?',(idInDb,))
                con.commit()
            self.LoadData(code=0)
        if delete_code==6:
            indices = self.ui.Position_Table.selectionModel().selectedRows()
            self.ui.Position_Table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Position_Table.removeRow(index.row())
                self.ui.Position_Table.selectionModel().clearCurrentIndex()
                self.ui.Position_Table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Position Where ID_Position = ?',(idInDb,))
                con.commit()
            self.LoadData(code=0)
            self.reload_position_combobox()
        if delete_code==7:
            indices = self.ui.Equipment_Table.selectionModel().selectedRows()
            self.ui.Equipment_Table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Equipment_Table.removeRow(index.row())
                self.ui.Equipment_Table.selectionModel().clearCurrentIndex()
                self.ui.Equipment_Table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Equipment Where ID_Equipment = ?',(idInDb,))
                con.commit()
            self.LoadData(code=0)
            self.Classrooms_Btn_clicked()
        if delete_code==8:
            indices = self.ui.Groups_Table.selectionModel().selectedRows()
            self.ui.Groups_Table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Groups_Table.removeRow(index.row())
                self.ui.Groups_Table.selectionModel().clearCurrentIndex()
                self.ui.Groups_Table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Groups Where ID_Groups = ?',(idInDb,))
                con.commit()
            self.LoadData(code=0)
        if delete_code==9:
            indices = self.ui.Teachers_Table.selectionModel().selectedRows()
            self.ui.Teachers_Table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Teachers_Table.removeRow(index.row())
                self.ui.Teachers_Table.selectionModel().clearCurrentIndex()
                self.ui.Teachers_Table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Teachers Where ID_Teachers = ?',(idInDb,))
                con.commit()
            self.LoadData(code=0)
        if delete_code==10:
            indices = self.ui.Classrooms_Table.selectionModel().selectedRows()
            self.ui.Classrooms_Table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Classrooms_Table.removeRow(index.row())
                self.ui.Classrooms_Table.selectionModel().clearCurrentIndex()
                self.ui.Classrooms_Table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Classrooms Where ID_Classrooms = ?',(idInDb,))
                con.commit()
            self.LoadData(code=0)
        if delete_code==11:
            indices = self.ui.Subject_Table.selectionModel().selectedRows()
            self.ui.Subject_Table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Subject_Table.removeRow(index.row())
                self.ui.Subject_Table.selectionModel().clearCurrentIndex()
                self.ui.Subject_Table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Subject Where ID_Subject = ?',(idInDb,))
                con.commit()
            self.LoadData(code=0)
        if delete_code==12:
            indices = self.ui.Teachers_subject_table.selectionModel().selectedRows()
            self.ui.Teachers_subject_table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Teachers_subject_table.removeRow(index.row())
                self.ui.Teachers_subject_table.selectionModel().clearCurrentIndex()
                self.ui.Teachers_subject_table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Teachers_Subject Where ID_Teachers_Subject = ?',(idInDb,))
                con.commit()
            self.LoadData(code=0)
        if delete_code==13:
            indices = self.ui.Semester_Groups_Subject_table.selectionModel().selectedRows()
            self.ui.Semester_Groups_Subject_table.selectionModel().currentIndex().row()
            for index in sorted(indices):
                idInDb = index.sibling(index.row(),0).data()
                self.ui.Semester_Groups_Subject_table.removeRow(index.row())
                self.ui.Semester_Groups_Subject_table.selectionModel().clearCurrentIndex()
                self.ui.Semester_Groups_Subject_table.selectionModel().clearSelection()
                cur.execute('DELETE FROM Semester_Groups_Subject Where ID_Semester_Groups_Subject = ?',(idInDb,))
                con.commit()
            self.LoadData(code=0)
        cur.close()
        con.close() 
        
    def Building_Institute_btn_clicked(self):
        self.ui.comboBox_Buildings_In_institute.clear()
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT Building_name, Building_placement FROM Building"
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                if column_number==0:
                    data_all=data
                else:
                    data_all=data_all+" ("+data+")"
                    self.ui.comboBox_Buildings_In_institute.addItem(data_all)
        cur.close() 
        con.close() 
        self.ui.stackedWidget.setCurrentIndex(0)

    def Department_Btn_clicked(self):
        self.ui.comboBox_Institute_In_Department.clear()
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT ID_Institute FROM Institute"
        query2 = "SELECT ID_Building FROM Institute"
        query3 = "SELECT Institute_Name FROM Institute WHERE ID_Institute = ?"
        query4 = "SELECT Building_name FROM Building WHERE ID_Building = ?"
        result = con.execute(query)
        list_of_Id_inst=[]
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                list_of_Id_inst.append(data)
        result2 = con.execute(query2)
        list_of_Id_building=[]
        for row_number , row_data in enumerate(result2):
            for column_number, data in enumerate(row_data):
                list_of_Id_building.append(data)
        for a in range (0,len(list_of_Id_building)):
            intitute=list_of_Id_inst[a]
            result3 = con.execute(query3,(intitute,))
            for row_number , row_data in enumerate(result3):
                for column_number, data in enumerate(row_data):
                    Instisute=data
            building=list_of_Id_building[a]
            result4 = con.execute(query4,(building,))
            for row_number , row_data in enumerate(result4):
                for column_number, data in enumerate(row_data):
                    building_name=data
            to_combobox=Instisute+" - "+building_name
            self.ui.comboBox_Institute_In_Department.addItem(to_combobox)
        cur.close() 
        con.close() 
        self.ui.stackedWidget.setCurrentIndex(1)

    def Semester_btn_clicked(self):
        global a
        a=[]
        self.ui.stackedWidget.setCurrentIndex(2)
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = ("SELECT ID_Semester, Semester_Name, Semester_Year FROM Semester")
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                if column_number==0:
                    data1=data
                if column_number==1:
                    data2=data
                if column_number==2:
                    data3=data2+" "+data
                    a.append(data1)
                    a.append(data3)
        global count_do
        count_do=len(a)
        count_do=int(count_do/2)
        v=1
        self.ui.Semester_And_Year_comboBox.clear()
        for b in range (0, count_do): 
            self.ui.Semester_And_Year_comboBox.addItem(a[v])
            v=v+2
        cur.close() 
        con.close()

    def Groups_Btn_clicked(self):
        self.ui.stackedWidget.setCurrentIndex(3)
        self.ui.Department_comboBox__in_groups.clear()
        data2=[]
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = ("SELECT Department_Name FROM Department")
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                data2.append(data)
        data2=set(data2) 
        data2=list(data2)
        for f in range (0, len(data2)):
            self.ui.Department_comboBox__in_groups.addItem(data2[f])
        cur.close() 
        con.close()

    def Teachers_Btn_clicked(self):
        self.ui.stackedWidget.setCurrentIndex(4)
        self.ui.Teachers_Name1.clear()
        self.ui.Teachers_Name2.clear()
        self.ui.Teachers_Name3.clear()
        self.ui.Teachers_comboBox_department.clear()
        self.ui.Teachers_comboBox_position.clear()
        data2=[]
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = ("SELECT Department_Name FROM Department")
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                data2.append(data)
        data2=set(data2) 
        data2=list(data2)
        for f in range (0, len(data2)):
            self.ui.Teachers_comboBox_department.addItem(data2[f])
        query = ("SELECT Position_name FROM Position")
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                self.ui.Teachers_comboBox_position.addItem(data)
        cur.close() 
        con.close()

    def reload_position_combobox(self):
        self.ui.Teachers_comboBox_position.clear()
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = ("SELECT Position_name FROM Position")
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                self.ui.Teachers_comboBox_position.addItem(data)
        cur.close() 
        con.close()

    def Classrooms_Btn_clicked(self):
        self.ui.stackedWidget.setCurrentIndex(5)
        self.ui.Classrooms_search_line.clear()
        self.ui.Classrooms_comboBox_3.clear()
        self.ui.Classrooms_comboBox_1.clear()
        self.ui.Classrooms_comboBox_2.clear()
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = ("SELECT Equipment_name FROM Equipment")
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                self.ui.Classrooms_comboBox_3.addItem(data)
        query = ("SELECT Building_name FROM Building")
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                self.ui.Classrooms_comboBox_1.addItem(data)
        cur.close() 
        con.close()

    def change_index_id_building(self):
        self.ui.Classrooms_comboBox_2.clear()
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        data_in_Classrooms_comboBox_1=self.ui.Classrooms_comboBox_1.itemText(self.ui.Classrooms_comboBox_1.currentIndex())
        query = ("SELECT Department_Name FROM Department WHERE Building_name LIKE ?")
        result = con.execute(query,(data_in_Classrooms_comboBox_1,))
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                self.ui.Classrooms_comboBox_2.addItem(data)
        cur.close() 
        con.close()

    def Subject_Btn_clicked(self):
        self.ui.stackedWidget.setCurrentIndex(6)
        self.ui.Subject_comboBox_department.clear()
        data2=[]
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = ("SELECT Department_Name FROM Department")
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                data2.append(data)
        data2=set(data2) 
        data2=list(data2)
        for f in range (0, len(data2)):
            self.ui.Subject_comboBox_department.addItem(data2[f])
        cur.close() 
        con.close()

    def on_selection_teachers(self):
        global pressed_teach, pressed_subj
        pressed_teach=1
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        global listOfVars_in_teach
        rows = self.ui.Teachers_Table_2.rowCount()
        cols = self.ui.Teachers_Table_2.columnCount()
        listOfVars_in_teach=[]
        listOfVars_in_teach.clear()
        self.clear_teachers()
        indices = self.ui.Teachers_Table_2.selectionModel().selectedRows()
        for index in sorted(indices):
            idInDb = index.sibling(index.row(),0).data()
            idInTb=index.row()
            getFromDb = cur.execute('SELECT * FROM Teachers Where ID_Teachers = ?',(idInDb,))
            listOfVars_in_teach_demo = getFromDb.fetchall()
            for row_number , row_data in enumerate(listOfVars_in_teach_demo):
                for column_number, data in enumerate(row_data):
                    listOfVars_in_teach.append(data)
            for n in range(0,cols):
                self.ui.Teachers_Table_2.item(idInTb,n).setBackground(QtGui.QColor(100,100,150))
        if pressed_subj==1:
            self.addInDatabase(add_code=12)
        cur.close() 
        con.close()

    def on_selection_in_subject(self):
        global pressed_subj, pressed_teach
        pressed_subj=1
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        global listofVars_in_subj
        listofVars_in_subj=[]
        rows = self.ui.Subject_Table_2.rowCount()
        cols = self.ui.Subject_Table_2.columnCount()
        listofVars_in_subj.clear()
        self.clear_subj()
        indices = self.ui.Subject_Table_2.selectionModel().selectedRows()
        for index in sorted(indices):
            idInDb = index.sibling(index.row(),0).data()
            idInTb=index.row()
            getFromDb = cur.execute('SELECT * FROM Subject Where ID_Subject = ?',(idInDb,))
            listofVars_in_subj_demo = getFromDb.fetchall()
            for row_number , row_data in enumerate(listofVars_in_subj_demo):
                for column_number, data in enumerate(row_data):
                    listofVars_in_subj.append(data)
            for n in range(0,cols):
                self.ui.Subject_Table_2.item(idInTb,n).setBackground(QtGui.QColor(100,100,150))
        if pressed_teach==1:
            self.addInDatabase(add_code=12)
        cur.close() 
        con.close()

    def clear_teachers(self):
        rows = self.ui.Teachers_Table_2.rowCount()
        cols = self.ui.Teachers_Table_2.columnCount()
        for a in range (0,rows):
            for b in range (0,cols):
                self.ui.Teachers_Table_2.item(a,b).setBackground(QtGui.QColor(255,255,255))

    def clear_subj(self):
        rows = self.ui.Subject_Table_2.rowCount()
        cols = self.ui.Subject_Table_2.columnCount()
        for a in range (0,rows):
            for b in range (0,cols):
                self.ui.Subject_Table_2.item(a,b).setBackground(QtGui.QColor(255,255,255))

    def clear_Semester(self):
        rows = self.ui.Semester_Table_2.rowCount()
        cols = self.ui.Semester_Table_2.columnCount()
        for a in range (0,rows):
            for b in range (0,cols):
                self.ui.Semester_Table_2.item(a,b).setBackground(QtGui.QColor(255,255,255))

    def clear_Groups(self):
        rows = self.ui.Groups_Table_2.rowCount()
        cols = self.ui.Groups_Table_2.columnCount()
        for a in range (0,rows):
            for b in range (0,cols):
                self.ui.Groups_Table_2.item(a,b).setBackground(QtGui.QColor(255,255,255))

    def clear_Subject(self):
        rows = self.ui.Subject_Table_3.rowCount()
        cols = self.ui.Subject_Table_3.columnCount()
        for a in range (0,rows):
            for b in range (0,cols):
                self.ui.Subject_Table_3.item(a,b).setBackground(QtGui.QColor(255,255,255))

    def sopost_teachers_subj(self):
        data2=[]
        data3=[]
        self.ui.stackedWidget.setCurrentIndex(7)
        self.ui.comboBox_depart.clear()
        self.ui.comboBox_depart_2.clear()
        self.ui.comboBox_Name.clear()
        self.ui.comboBox_depart.addItem("Все кафедры")
        self.ui.comboBox_depart_2.addItem("Все кафедры")
        self.ui.comboBox_Name.addItem("Все преподаватели")
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT Department_Name FROM Department"
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                data2.append(data)
        data2=set(data2)
        data2=list(data2)
        for a in range (0,len(data2)):
            now=data2[a]
            self.ui.comboBox_depart.addItem(now)
            self.ui.comboBox_depart_2.addItem(now)
        query = "SELECT Teachers_Name_1 FROM Teachers"
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                data3.append(data)
        data3=set(data3)
        data3=list(data3)
        data3.sort()
        for a in range (0,len(data3)):
            now=data3[a]
            self.ui.comboBox_Name.addItem(now)
        cur.close() 
        con.close()
    
    def comboBox_depart_change(self):
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT * FROM Teachers WHERE Department_Name LIKE ?"
        query2 = "SELECT * FROM Teachers WHERE Teachers_Name_1 LIKE ?"
        query3 = "SELECT * FROM Teachers WHERE Department_Name LIKE ? AND Teachers_Name_1 LIKE ?"
        kafedra=self.ui.comboBox_depart.itemText(self.ui.comboBox_depart.currentIndex())
        Name=self.ui.comboBox_Name.itemText(self.ui.comboBox_Name.currentIndex())
        if kafedra=="Все кафедры":
            if Name=="Все преподаватели":
                result = con.execute("SELECT * FROM Teachers")
            else:
                result = con.execute(query2,(Name,))
        else:
            if Name=="Все преподаватели":
                result = con.execute(query,(kafedra,))
            else:
                result = con.execute(query3,(kafedra,Name,))            
        self.ui.Teachers_Table_2.setRowCount(0)
        for row_number , row_data in enumerate(result):
            self.ui.Teachers_Table_2.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.ui.Teachers_Table_2.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data))) 
        self.ui.Teachers_Table_2.resizeColumnsToContents()
        cur.close() 
        con.close()
        
    def comboBox_depart_2_change(self):
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT * FROM Subject WHERE Department_Name LIKE ?"
        aga=self.ui.comboBox_depart_2.itemText(self.ui.comboBox_depart_2.currentIndex())
        if aga=="Все кафедры":
            result = con.execute("SELECT * FROM Subject")
        else:
            result = con.execute(query,(self.ui.comboBox_depart_2.itemText(self.ui.comboBox_depart_2.currentIndex()),))
        self.ui.Subject_Table_2.setRowCount(0)
        for row_number , row_data in enumerate(result):
            self.ui.Subject_Table_2.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.ui.Subject_Table_2.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data))) 
        self.ui.Subject_Table_2.resizeColumnsToContents() 
        cur.close() 
        con.close()

    def comboBox_Name_change(self):
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT * FROM Teachers WHERE Teachers_Name_1 LIKE ?"
        query2 = "SELECT * FROM Teachers WHERE Department_Name LIKE ?"
        query3 = "SELECT * FROM Teachers WHERE Department_Name LIKE ? AND Teachers_Name_1 LIKE ?"
        Name=self.ui.comboBox_Name.itemText(self.ui.comboBox_Name.currentIndex())
        kafedra=self.ui.comboBox_depart.itemText(self.ui.comboBox_depart.currentIndex())
        if Name=="Все преподаватели":
            if kafedra=="Все кафедры":
                result = con.execute("SELECT * FROM Teachers")
            else:
                result = con.execute(query2,(kafedra,))
        else:
            if kafedra=="Все кафедры":
                result = con.execute(query,(Name,))
            else:
                result = con.execute(query3,(kafedra,Name,))
        self.ui.Teachers_Table_2.setRowCount(0)
        for row_number , row_data in enumerate(result):
            self.ui.Teachers_Table_2.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.ui.Teachers_Table_2.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        self.ui.Teachers_Table_2.resizeColumnsToContents() 
        cur.close() 
        con.close()

    def sopost_semestr_groups_subj(self):
        data2=[]
        self.ui.stackedWidget.setCurrentIndex(8)
        self.ui.comboBox_depart_3.clear()
        self.ui.comboBox_depart_4.clear()
        self.ui.comboBox_depart_3.addItem("Все кафедры")
        self.ui.comboBox_depart_4.addItem("Все кафедры")
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT Department_Name FROM Department"
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                data2.append(data)
        data2=set(data2)
        data2=list(data2)
        for a in range (0,len(data2)):
            now=data2[a]
            self.ui.comboBox_depart_3.addItem(now)
            self.ui.comboBox_depart_4.addItem(now)
        cur.close() 
        con.close()
    
    def on_selection_Semester_Table_2(self):
        global pressed_semestr, pressed_groups, pressed_subject
        pressed_semestr=1
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        global listOfVars_in_semestr
        rows = self.ui.Semester_Table_2.rowCount()
        cols = self.ui.Semester_Table_2.columnCount()
        listOfVars_in_semestr=[]
        listOfVars_in_semestr.clear()
        self.clear_Semester()
        indices = self.ui.Semester_Table_2.selectionModel().selectedRows()
        for index in sorted(indices):
            idInDb=index.sibling(index.row(),0).data()
            idInTb=index.row()
            getFromDb = cur.execute('SELECT * FROM Semester Where ID_Semester = ?',(idInDb,))
            listOfVars_in_semestr_demo = getFromDb.fetchall()
            for row_number , row_data in enumerate(listOfVars_in_semestr_demo):
                for column_number, data in enumerate(row_data):
                    listOfVars_in_semestr.append(data)
            for n in range(0,cols):
                self.ui.Semester_Table_2.item(idInTb,n).setBackground(QtGui.QColor(100,100,150))
        if pressed_groups==1:
            if pressed_subject==1:
                self.addInDatabase(add_code=13)
        cur.close() 
        con.close()

    def on_selection_Groups_Table_2(self):
        global pressed_semestr, pressed_groups, pressed_subject
        pressed_groups=1
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        global listOfVars_in_groups
        rows = self.ui.Groups_Table_2.rowCount()
        cols = self.ui.Groups_Table_2.columnCount()
        listOfVars_in_groups=[]
        listOfVars_in_groups.clear()
        self.clear_Groups()
        indices = self.ui.Groups_Table_2.selectionModel().selectedRows()
        for index in sorted(indices):
            idInDb = index.sibling(index.row(),0).data()
            idInTb=index.row()
            getFromDb = cur.execute('SELECT * FROM Groups Where ID_Groups = ?',(idInDb,))
            listOfVars_in_groups_demo = getFromDb.fetchall()
            for row_number , row_data in enumerate(listOfVars_in_groups_demo):
                for column_number, data in enumerate(row_data):
                    listOfVars_in_groups.append(data)
            for n in range(0,cols):
                self.ui.Groups_Table_2.item(idInTb,n).setBackground(QtGui.QColor(100,100,150))
        if pressed_semestr==1:
            if pressed_subject==1:
                self.addInDatabase(add_code=13)
        cur.close() 
        con.close()
    
    def on_selection_Subject_Table_3(self):
        global pressed_semestr, pressed_groups, pressed_subject
        pressed_subject=1
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        global listOfVars_in_subject
        rows = self.ui.Subject_Table_3.rowCount()
        cols = self.ui.Subject_Table_3.columnCount()
        listOfVars_in_subject=[]
        listOfVars_in_subject.clear()
        self.clear_Subject()
        indices = self.ui.Subject_Table_3.selectionModel().selectedRows()
        for index in sorted(indices):
            idInDb = index.sibling(index.row(),0).data()
            idInTb=index.row()
            getFromDb = cur.execute('SELECT * FROM Subject Where ID_Subject = ?',(idInDb,))
            listOfVars_in_subject_demo = getFromDb.fetchall()
            for row_number , row_data in enumerate(listOfVars_in_subject_demo):
                for column_number, data in enumerate(row_data):
                    listOfVars_in_subject.append(data)
            for n in range(0,cols):
                self.ui.Subject_Table_3.item(idInTb,n).setBackground(QtGui.QColor(100,100,150))
        if pressed_semestr==1:
            if pressed_groups==1:
                self.addInDatabase(add_code=13)
        cur.close() 
        con.close()

    def check_choise(self):
        global pressed_teach, pressed_subj
        pressed_teach=0
        pressed_subj=0
    
    def check_choise_2(self):
        global pressed_semestr, pressed_groups, pressed_subject
        pressed_semestr=0
        pressed_groups=0
        pressed_subject=0

    def comboBox_depart_3_change(self):
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT * FROM Groups WHERE Department_Name LIKE ?"
        query2 = "SELECT * FROM Groups WHERE Groups_Type LIKE ?"
        query3 = "SELECT * FROM Groups WHERE Department_Name LIKE ? AND Groups_Type LIKE ?"
        kafedra=self.ui.comboBox_depart_3.itemText(self.ui.comboBox_depart_3.currentIndex())
        Type=self.ui.comboBox_type_group.itemText(self.ui.comboBox_type_group.currentIndex())
        if kafedra=="Все кафедры":
            if Type=="Любой":
                result = con.execute("SELECT * FROM Groups")
            else:
                result = con.execute(query2,(Type,))
        else:
            if Type=="Любой":
                result = con.execute(query,(kafedra,))
            else:
                result = con.execute(query3,(kafedra,Type,))            
        self.ui.Groups_Table_2.setRowCount(0)
        for row_number , row_data in enumerate(result):
            self.ui.Groups_Table_2.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.ui.Groups_Table_2.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data))) 
        self.ui.Groups_Table_2.resizeColumnsToContents()
        cur.close() 
        con.close()

    def comboBox_type_group_change(self):
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT * FROM Groups WHERE Groups_Type LIKE ?"
        query2 = "SELECT * FROM Groups WHERE Department_Name LIKE ?"
        query3 = "SELECT * FROM Groups WHERE Department_Name LIKE ? AND Groups_Type LIKE ?"
        Type=self.ui.comboBox_type_group.itemText(self.ui.comboBox_type_group.currentIndex())
        kafedra=self.ui.comboBox_depart_3.itemText(self.ui.comboBox_depart_3.currentIndex())
        if Type=="Любой":
            if kafedra=="Все кафедры":
                result = con.execute("SELECT * FROM Groups")
            else:
                result = con.execute(query2,(kafedra,))
        else:
            if kafedra=="Все кафедры":
                result = con.execute(query,(Type,))
            else:
                result = con.execute(query3,(kafedra,Type,))
        self.ui.Groups_Table_2.setRowCount(0)
        for row_number , row_data in enumerate(result):
            self.ui.Groups_Table_2.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.ui.Groups_Table_2.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data))) 
        self.ui.Groups_Table_2.resizeColumnsToContents()
        cur.close() 
        con.close()

    def comboBox_depart_4_change(self):
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT * FROM Subject WHERE Department_Name LIKE ?"
        aga=self.ui.comboBox_depart_4.itemText(self.ui.comboBox_depart_4.currentIndex())
        if aga=="Все кафедры":
            result = con.execute("SELECT * FROM Subject")
        else:
            result = con.execute(query,(self.ui.comboBox_depart_4.itemText(self.ui.comboBox_depart_4.currentIndex()),))
        self.ui.Subject_Table_3.setRowCount(0)
        for row_number , row_data in enumerate(result):
            self.ui.Subject_Table_3.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.ui.Subject_Table_3.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data))) 
        self.ui.Subject_Table_3.resizeColumnsToContents()
        cur.close() 
        con.close()
    
    def set_combobox(self):
        self.ui.stackedWidget.setCurrentIndex(9)
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        self.ui.semestr_combobox.clear()
        query = "SELECT Semester_Name, Semester_Year FROM Semester"
        result = con.execute(query)
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                if column_number==0:
                    data_all=data
                else:
                    data_all=data_all+" ("+data+")"
                    self.ui.semestr_combobox.addItem(data_all) 
        self.ui.comboBox_teachers_done.clear()
        self.ui.comboBox_teachers_done.addItem("Все преподаватели")
        data2=[]
        query2 = ("SELECT Teachers_Name_1, Teachers_Name_2, Teachers_Name_3 FROM Done")
        result2 = con.execute(query2)
        for row_number , row_data in enumerate(result2):
            for column_number, data in enumerate(row_data):
                data2.append(data)
        score=0
        listtoset=[]
        for f in range (0, len(data2)):
            if score==0:
                datatoadd=(data2[f])
                score=score+1
            elif score==1:
                datatoadd=datatoadd+" "+(data2[f])
                score=score+1
            elif score==2:
                datatoadd=datatoadd+" "+(data2[f])
                listtoset.append(datatoadd)
                score=0
        listtoset=list(set(listtoset))
        for c in range (0, len(listtoset)):
            self.ui.comboBox_teachers_done.addItem(listtoset[c])
        self.ui.comboBox_groups_done.clear()
        self.ui.comboBox_groups_done.addItem("Все группы")
        data3=[]
        query3 = ("SELECT Groups_name FROM Done")
        result3 = con.execute(query3)
        for row_number , row_data in enumerate(result3):
            for column_number, data in enumerate(row_data):
                data3.append(data)
        data3=list(set(data3))
        for f in range (0, len(data3)):
            self.ui.comboBox_groups_done.addItem(data3[f])
        self.ui.comboBox_class_done.clear()
        self.ui.comboBox_class_done.addItem("Все аудитории")
        data4=[]
        query4 = ("SELECT Classrooms_Name FROM Done")
        result4 = con.execute(query4)
        for row_number , row_data in enumerate(result4):
            for column_number, data in enumerate(row_data):
                data4.append(data)
        data4=list(set(data4))
        for f in range (0, len(data4)):
            self.ui.comboBox_class_done.addItem(data4[f])
        cur.close()
        con.close()

    def start(self):
        self.drop_table()
        self.completed=0
        self.ui.lineedit_score.setText("Производятся вычисления...")
        data_id_groups=[]
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        indices=self.ui.semestr_combobox.currentIndex()
        id_in_semestr=int(self.ui.Semester_Table_2.item(indices, 0).text())
        query = "SELECT ID_Groups FROM Semester_Groups_Subject WHERE ID_Semester = ?"
        result = con.execute(query,(id_in_semestr,))
        for row_number , row_data in enumerate(result):
            for column_number, data in enumerate(row_data):
                data_id_groups.append(data)
        data_id_groups=list(set(data_id_groups))
        query2 = "SELECT ID_Subject FROM Semester_Groups_Subject WHERE ID_Groups = ?"
        query3 = "SELECT ID_Teachers FROM Teachers_Subject WHERE ID_Subject = ?"
        query4 = "SELECT Department_Name FROM Teachers_Subject WHERE ID_Subject = ?"
        query5 = "SELECT ID_Classrooms FROM Classrooms WHERE Department_Name = ?"
        query6 = "SELECT Teachers_Name_1, Teachers_Name_2, Teachers_Name_3 FROM Teachers WHERE ID_Teachers = ?"
        query7 = "SELECT Groups_name FROM Groups WHERE ID_Groups = ?"
        query8 = "SELECT Classrooms_Name FROM Classrooms WHERE ID_Classrooms = ?"
        query9 = "SELECT Subject_Name, Subject_type FROM Subject WHERE ID_Subject = ?"
        query10 = "SELECT Number_of_day FROM Done WHERE ID_Teachers = ?"
        query11 = "SELECT Number_of_day FROM Done WHERE ID_Classrooms = ?"
        query12 = "SELECT Number_of_day FROM Done WHERE ID_Groups = ?"
        for a in range(0, len(data_id_groups)):
            data_id_subject=[]
            group_now=data_id_groups[a]
            result2 = con.execute(query2,(group_now,))
            for row_number , row_data in enumerate(result2):
                for column_number, data in enumerate(row_data):
                    data_id_subject.append(data)
            for b in range (0,len(data_id_subject)):
                data_teachers=[]
                subject_id_now=data_id_subject[b]
                result3 = con.execute(query3,(subject_id_now,))
                for row_number , row_data in enumerate(result3):
                    for column_number, data in enumerate(row_data):
                        data_teachers.append(data)
                result4 = con.execute(query4,(subject_id_now,))
                for row_number , row_data in enumerate(result4):
                    for column_number, data in enumerate(row_data):
                        department_name_now=data
                id_classrooms_for_this_department=[]
                result5 = con.execute(query5,(department_name_now,))
                for row_number , row_data in enumerate(result5):
                    for column_number, data in enumerate(row_data):
                        id_classrooms_for_this_department.append(data)
                list_zanyat_prepod=[]
                list_zanyata_gruppa=[]
                id_teachers_done=data_teachers[0]
                number_of_day=1
                result10 = con.execute(query10,(id_teachers_done,))
                for row_number , row_data in enumerate(result10):
                    for column_number, data in enumerate(row_data):
                        list_zanyat_prepod.append(data)
                list_zanyat_prepod=list(set(list_zanyat_prepod))
                result12 = con.execute(query12,(group_now,))
                for row_number , row_data in enumerate(result12):
                    for column_number, data in enumerate(row_data):
                        list_zanyata_gruppa.append(data)
                list_zanyata_gruppa=list(set(list_zanyata_gruppa))
                for d in range(0,25):
                    for i in range(0,len(list_zanyat_prepod)):
                        check=list_zanyat_prepod[i]
                        if check == number_of_day:
                            number_of_day+=1
                    for i in range(0,len(list_zanyata_gruppa)):
                        check=list_zanyata_gruppa[i]
                        if check == number_of_day:
                            number_of_day+=1
                fio=[]
                result6 = con.execute(query6,(id_teachers_done,))
                for row_number , row_data in enumerate(result6):
                    for column_number, data in enumerate(row_data):
                        fio.append(data)
                if 1 <= number_of_day <=4:
                    den_nedely="Понедельник"
                elif 5 <= number_of_day <=8:
                    den_nedely="Вторник"
                elif 9 <= number_of_day <=12:
                    den_nedely="Среда"
                elif 13 <= number_of_day <=16:
                    den_nedely="Четверг"
                elif 17 <= number_of_day <=20:
                    den_nedely="Пятница"
                elif 21 <= number_of_day <=24:
                    den_nedely="Суббота"
                elif 25 <= number_of_day <=99999:
                    den_nedely="Ошибка"
                    number_para="Ошибка"
                if number_of_day in (1,5,9,13,17,21):
                    number_para=1
                elif number_of_day in (2,6,10,14,18,22):
                    number_para=2
                elif number_of_day in (3,7,11,15,19,23):
                    number_para=3
                elif number_of_day in (4,8,12,16,20,24):
                    number_para=4
                its_in=1
                for m in range (0, len(id_classrooms_for_this_department)):
                    if its_in<=0:
                        its_in=0
                        break
                    elif its_in>=2:
                        list_zanyata_auditoriya=[]
                        classroom_done=id_classrooms_for_this_department[m]
                        result11 = con.execute(query11,(classroom_done,))
                        for row_number , row_data in enumerate(result11):
                            for column_number, data in enumerate(row_data):
                                list_zanyata_auditoriya.append(data)
                        list_zanyata_auditoriya=list(set(list_zanyata_auditoriya))
                        if not list_zanyata_auditoriya:
                            break 
                        else:
                            number_of_day_check=[]
                            number_of_day_check.append(number_of_day)
                            for i in list_zanyata_auditoriya:
                                for j in number_of_day_check:
                                    if i==j:
                                        its_in+=1
                                    else:
                                        its_in+=0
                    else:
                        list_zanyata_auditoriya=[]
                        classroom_done=id_classrooms_for_this_department[m]
                        result11 = con.execute(query11,(classroom_done,))
                        for row_number , row_data in enumerate(result11):
                            for column_number, data in enumerate(row_data):
                                list_zanyata_auditoriya.append(data)
                        list_zanyata_auditoriya=list(set(list_zanyata_auditoriya))
                        minus_kolvo_dney=len(list_zanyata_auditoriya)
                        minus_kolvo_dney+=1
                        if not list_zanyata_auditoriya:
                            break 
                        else:
                            number_of_day_check=[]
                            number_of_day_check.append(number_of_day)
                            for i in list_zanyata_auditoriya:
                                for j in number_of_day_check:
                                    if i==j:
                                        its_in+=2
                                    else:
                                        its_in+=1
                            its_in=its_in-minus_kolvo_dney
                result7 = con.execute(query7,(group_now,))
                for row_number , row_data in enumerate(result7):
                    for column_number, data in enumerate(row_data):
                        group_name=data
                result8 = con.execute(query8,(classroom_done,))
                for row_number , row_data in enumerate(result8):
                    for column_number, data in enumerate(row_data):
                        classroom_name=data
                subj_data=[]
                result9 = con.execute(query9,(subject_id_now,))
                for row_number , row_data in enumerate(result9):
                    for column_number, data in enumerate(row_data):
                        subj_data.append(data)
                subj_name=subj_data[0]
                subj_type=subj_data[1]
                list_to_add=[]
                list_to_add.append(id_teachers_done)
                list_to_add.extend(fio)
                list_to_add.append(number_of_day)
                list_to_add.append(den_nedely)
                list_to_add.append(number_para)
                list_to_add.append(group_now)
                list_to_add.append(group_name)
                list_to_add.append(classroom_done)
                list_to_add.append(classroom_name)
                list_to_add.append(subject_id_now)
                list_to_add.append(subj_name)
                list_to_add.append(subj_type)
                cur.execute('INSERT INTO Done VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)',list_to_add)
                con.commit() 
                self.LoadData(code=14)
        while self.completed<100:
            self.completed +=1
            self.ui.progressBar.setValue(self.completed)
        self.ui.lineedit_score.setText("Расписание успешно составлено")
        cur.close() 
        con.close()

    def search_in_teachers(self): 
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = """SELECT * FROM Teachers WHERE ? IN (
        Teachers_Name_1,Teachers_Name_2,Teachers_Name_3,
        Teachers_Name_1 || ' ' || Teachers_Name_2 || ' ' || Teachers_Name_3, /*ФИО*/
        Teachers_Name_1 || ' ' || Teachers_Name_3 || ' ' || Teachers_Name_2, /*ФОИ*/
        Teachers_Name_1 || ' ' || Teachers_Name_2, /*ФИ*/
        Teachers_Name_1 || ' ' || Teachers_Name_3, /*ФО*/
        Teachers_Name_2 || ' ' || Teachers_Name_3 || ' ' || Teachers_Name_1, /*ИОФ*/
        Teachers_Name_2 || ' ' || Teachers_Name_1 || ' ' || Teachers_Name_3, /*ИФО*/
        Teachers_Name_2 || ' ' || Teachers_Name_3, /*ИО*/
        Teachers_Name_2 || ' ' || Teachers_Name_1, /*ИФ*/
        Teachers_Name_3 || ' ' || Teachers_Name_1 || ' ' || Teachers_Name_2, /*ОФИ*/
        Teachers_Name_3 || ' ' || Teachers_Name_2 || ' ' || Teachers_Name_1, /*ОИФ*/
        Teachers_Name_3 || ' ' || Teachers_Name_1, /*ОФ*/
        Teachers_Name_3 || ' ' || Teachers_Name_2 /*ОИ*/
        )
        """
        if self.ui.Teachers_search_line.text() == '':
            self.LoadData(code=0)
        else:
            result = con.execute(query,(self.ui.Teachers_search_line.text(),))
            self.ui.Teachers_Table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Teachers_Table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Teachers_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        cur.close() 
        con.close()

    def search_in_class(self): 
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT * FROM Classrooms WHERE Classrooms_Name LIKE ?"
        if self.ui.Classrooms_search_line.text() == '':
            self.LoadData(code=0)
        else:
            result = con.execute(query,(self.ui.Classrooms_search_line.text(),))
            self.ui.Classrooms_Table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Classrooms_Table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Classrooms_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        cur.close() 
        con.close()
    
    def search_in_subj(self):
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT * FROM Subject WHERE Subject_Name LIKE ?"
        if self.ui.Subject_search_line.text() == '':
            self.LoadData(code=0)
        else:
            result = con.execute(query,(self.ui.Subject_search_line.text(),))
            self.ui.Subject_Table.setRowCount(0)
            for row_number , row_data in enumerate(result):
                self.ui.Subject_Table.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.ui.Subject_Table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        cur.close() 
        con.close()

    def drop_table(self):
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        cur.execute("DROP TABLE Done")
        con.commit()
        cur.close() 
        con.close()
        self.create_Tables()
        self.LoadData(code=0)
    
    def comboBox_teachers_done_def(self):
        self.ui.comboBox_groups_done.setCurrentIndex(0)
        self.ui.comboBox_class_done.setCurrentIndex(0)
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = """SELECT * FROM Done WHERE ? IN (
        Teachers_Name_1,Teachers_Name_2,Teachers_Name_3,
        Teachers_Name_1 || ' ' || Teachers_Name_2 || ' ' || Teachers_Name_3 /*ФИО*/
        )
        """
        aga=self.ui.comboBox_teachers_done.itemText(self.ui.comboBox_teachers_done.currentIndex())
        if aga=="Все преподаватели":
            result = con.execute("SELECT * FROM Done")
        else:
            result = con.execute(query,(self.ui.comboBox_teachers_done.itemText(self.ui.comboBox_teachers_done.currentIndex()),))
        self.ui.Done_table.setRowCount(0)
        for row_number , row_data in enumerate(result):
            self.ui.Done_table.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.ui.Done_table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data))) 
        self.ui.Done_table.resizeColumnsToContents()
        cur.close() 
        con.close()

    def comboBox_groups_done_def(self):
        self.ui.comboBox_teachers_done.setCurrentIndex(0)
        self.ui.comboBox_class_done.setCurrentIndex(0)
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT * FROM Done WHERE Groups_name LIKE ?"
        aga=self.ui.comboBox_groups_done.itemText(self.ui.comboBox_groups_done.currentIndex())
        if aga=="Все группы":
            result = con.execute("SELECT * FROM Done")
        else:
            result = con.execute(query,(self.ui.comboBox_groups_done.itemText(self.ui.comboBox_groups_done.currentIndex()),))
        self.ui.Done_table.setRowCount(0)
        for row_number , row_data in enumerate(result):
            self.ui.Done_table.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.ui.Done_table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data))) 
        self.ui.Done_table.resizeColumnsToContents()
        cur.close() 
        con.close()
    
    def comboBox_class_done_def(self):
        self.ui.comboBox_teachers_done.setCurrentIndex(0)
        self.ui.comboBox_groups_done.setCurrentIndex(0)
        con = sqlite3.connect("./mydatabase.db")
        con.execute("PRAGMA foreign_keys = 1")
        cur = con.cursor()
        query = "SELECT * FROM Done WHERE Classrooms_Name LIKE ?"
        aga=self.ui.comboBox_class_done.itemText(self.ui.comboBox_class_done.currentIndex())
        if aga=="Все аудитории":
            result = con.execute("SELECT * FROM Done")
        else:
            result = con.execute(query,(self.ui.comboBox_class_done.itemText(self.ui.comboBox_class_done.currentIndex()),))
        self.ui.Done_table.setRowCount(0)
        for row_number , row_data in enumerate(result):
            self.ui.Done_table.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.ui.Done_table.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data))) 
        self.ui.Done_table.resizeColumnsToContents()
        cur.close() 
        con.close()

    def AboutPythonDef(self):
        Dialog_python = QtWidgets.QDialog()
        ui = Ui_Dialog_python()
        ui.setupUi(Dialog_python)
        Dialog_python.show()
        Dialog_python.exec_()

    def Aboutautors(self):
        about_autors = QtWidgets.QDialog()
        ui = Ui_about_autors()
        ui.setupUi(about_autors)
        about_autors.show()
        about_autors.exec_()
    
    def spbgltu(self):
        spbgltu = QtWidgets.QDialog()
        ui = Ui_spbgltu()
        ui.setupUi(spbgltu)
        spbgltu.show()
        spbgltu.exec_()
    
    def save(self):
        wb = Workbook()
        wb.create_sheet(title = 'Первый лист', index = 0)
        sheet = wb['Первый лист']
        teext="Фамилия"
        cell = sheet.cell(row = 1, column = 1) 
        cell.value=teext
        teext="Имя"
        cell = sheet.cell(row = 1, column = 2) 
        cell.value=teext
        teext="Отчество"
        cell = sheet.cell(row = 1, column = 3)
        cell.value=teext
        teext="День"
        cell = sheet.cell(row = 1, column = 4) 
        cell.value=teext
        teext="День недели"
        cell = sheet.cell(row = 1, column = 5) 
        cell.value=teext
        teext="Номер пары"
        cell = sheet.cell(row = 1, column = 6) 
        cell.value=teext
        teext="Группа"
        cell = sheet.cell(row = 1, column = 7) 
        cell.value=teext
        teext="Аудитория"
        cell = sheet.cell(row = 1, column = 8) 
        cell.value=teext
        teext="Предмет"
        cell = sheet.cell(row = 1, column = 9) 
        cell.value=teext
        teext="Тип предмета"
        cell = sheet.cell(row = 1, column = 10) 
        cell.value=teext
        row_for_table = -1
        row=1
        stolbcov=self.ui.Done_table.columnCount()
        strok=self.ui.Done_table.rowCount()
        for x in range(self.ui.Done_table.rowCount()):
            column=1
            col_for_table = 0
            row+=1
            row_for_table += 1
            for i in range(self.ui.Done_table.columnCount()):
                try:             
                    if col_for_table==0:
                        col_for_table += 1
                    if col_for_table==7:
                        col_for_table += 1
                    if col_for_table==9:
                        col_for_table += 1
                    if col_for_table==11:
                        col_for_table += 1
                    else:
                        teext = str(self.ui.Done_table.item(row_for_table, col_for_table).text())
                        cell = sheet.cell(row = row, column = column)
                        cell.value=teext 
                        col_for_table += 1
                        column += 1
                except AttributeError:
                    None
        wb.save('Расписание.xlsx')

if __name__=='__main__':
    app = QtWidgets.QApplication(sys.argv)
    myapp = MyWin() 
    myapp.show() 
    sys.exit(app.exec_())